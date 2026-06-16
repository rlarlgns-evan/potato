#!/usr/bin/env python3
"""엑셀 시군구 코드 → data/gangwon_sigungu_codes.json (강원만)."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "gangwon_sigungu_codes.json"
DEFAULT_XLSX = (
    Path.home()
    / "Downloads"
    / "TourAPI_Guide_(중심관광지)v4.1"
    / "한국관광공사_TourAPI_관광지_시군구_코드정보_v1.0.xlsx"
)


def import_gangwon_codes(xlsx_path: Path) -> dict:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl 필요: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {str(h): i for i, h in enumerate(header)}

    regions: dict[str, dict] = {}
    area_cd = None
    area_nm = None

    for row in rows[1:]:
        if not row or row[idx["areaNm"]] is None:
            continue
        name = str(row[idx["areaNm"]])
        if "강원" not in name:
            continue
        area_cd = int(row[idx["areaCd"]])
        area_nm = name
        signgu_nm = str(row[idx["sigunguNm"]]).strip()
        regions[signgu_nm] = {
            "areaCd": area_cd,
            "areaNm": area_nm,
            "signguCd": int(row[idx["sigunguCd"]]),
            "signguNm": signgu_nm,
        }

    if not regions:
        raise SystemExit("엑셀에서 강원 시군구 코드를 찾지 못했습니다.")

    return {
        "source": "한국관광공사_TourAPI_관광지_시군구_코드정보_v1.0",
        "areaCd": area_cd,
        "areaNm": area_nm,
        "regions": regions,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"파일 없음: {args.xlsx}", file=sys.stderr)
        return 1

    payload = import_gangwon_codes(args.xlsx)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUT} ({len(payload['regions'])} regions)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
